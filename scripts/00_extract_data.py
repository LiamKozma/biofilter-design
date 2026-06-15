#!/usr/bin/env python3
"""Extract tidy, analysis-ready data from the original BCHE 3420 spreadsheet.

The source workbook (`Kozma design project 1.xlsx`) is a single-sheet working
scratchpad: six tables are packed side by side, interleaved with intermediate
regression columns and Excel error cells (#DIV/0!, #NUM!). This script reads the
*canonical* measurements only -- the bench-scale concentration profiles and the
two design tables -- and writes them as flat CSVs with explicit units, so every
downstream stage consumes clean, documented inputs rather than re-parsing the
spreadsheet.

Bench-scale operating conditions (from the report, Appendix A, Table 1):
    Q   = 4 L/min volumetric flow            (= 6.6667e-5 m^3/s)
    D   = 0.10 m column diameter             (A = pi D^2 / 4)
    L   = 0.50 m packed-bed length
    ports at 0.00, 0.15, 0.25, 0.35, 0.50 m from the inlet

The empty-bed residence time to a given port is t = A * x / Q -- this is the
independent variable for the integral-method kinetics, reproducing the report's
`t = V/Q` column.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC_XLSX = ROOT / "data" / "raw" / "Kozma design project 1.xlsx"
OUT = ROOT / "data" / "tidy"
OUT.mkdir(parents=True, exist_ok=True)

# Bench-scale operating conditions.
Q_LMIN = 4.0                      # L/min
Q_M3MIN = Q_LMIN * 1e-3           # m^3/min
Q_M3S = Q_M3MIN / 60.0            # m^3/s
DIAM_M = 0.10                     # m
AREA_M2 = 3.141592653589793 * DIAM_M**2 / 4.0
BED_LEN_M = 0.50                  # m

# Compound columns within each concentration block (offset from "Position" col).
COMPOUNDS = ["CH3SH", "3-MB", "2-MB", "Hexanal", "DMDS"]
PORT_LABELS = ["Inlet", "Port 1", "Port 2", "Port 3", "Outlet"]


def residence_time_min(position_m: float) -> float:
    """Empty-bed residence time (min) for a port at `position_m` from inlet."""
    return AREA_M2 * position_m / Q_M3MIN


def find_profile_blocks(ws):
    """Locate concentration blocks by their header signature.

    A block header is a row whose column C reads 'Position (m)' and whose column
    B holds the run date. The five data rows follow immediately below it.
    """
    blocks = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 3).value == "Position (m)":          # column C
            date_cell = ws.cell(r, 2).value                # column B
            if isinstance(date_cell, dt.datetime):
                blocks.append((r, date_cell.date()))
    return blocks


def extract_profiles(ws):
    rows = []
    for header_row, run_date in find_profile_blocks(ws):
        for i, port in enumerate(PORT_LABELS):
            r = header_row + 1 + i
            position = ws.cell(r, 3).value                 # column C
            if position is None:
                continue
            for j, compound in enumerate(COMPOUNDS):
                val = ws.cell(r, 4 + j).value              # columns D..H
                if isinstance(val, (int, float)):
                    rows.append(
                        {
                            "run_date": run_date.isoformat(),
                            "port": port,
                            "position_m": float(position),
                            "residence_time_min": residence_time_min(float(position)),
                            "compound": compound,
                            "conc_ppmv": float(val),
                        }
                    )
    return rows


def extract_influent(ws):
    """Table 2: VOC composition of the industrial gas stream to be treated.

    Lives in columns AF/AG/AH (32/33/34). Two grab samples; non-numeric entries
    ('P' present, 'ND' not detected) are preserved as strings.
    """
    rows = []
    for r in range(6, 19):                                 # Methanethiol..Total
        name = ws.cell(r, 32).value                        # AF
        if not name or name in ("Total",):
            if name == "Total":
                pass  # keep the total row too
            else:
                continue
        s1 = ws.cell(r, 33).value                          # AG
        s2 = ws.cell(r, 34).value                          # AH
        rows.append({"compound": name, "sample_1_ppmv": s1, "sample_2_ppmv": s2})
    return rows


def extract_media(ws):
    """Table 3: physical properties of the candidate biocatalyst media."""
    rows = []
    for r in range(5, 10):                                 # property rows
        prop = ws.cell(r, 37).value                        # AK
        biosorbens = ws.cell(r, 38).value                  # AL
        compost = ws.cell(r, 39).value                     # AM
        if prop:
            rows.append(
                {"property": prop, "BIOSORBENS": biosorbens, "compost": compost}
            )
    return rows


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
    import csv

    with path.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {len(rows):3d} rows -> {path.relative_to(ROOT)}")


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["Sheet1"]

    profiles = extract_profiles(ws)
    write_csv(
        OUT / "profiles.csv",
        profiles,
        ["run_date", "port", "position_m", "residence_time_min", "compound", "conc_ppmv"],
    )

    influent = extract_influent(ws)
    write_csv(OUT / "influent_table2.csv", influent,
              ["compound", "sample_1_ppmv", "sample_2_ppmv"])

    media = extract_media(ws)
    write_csv(OUT / "media_table3.csv", media,
              ["property", "BIOSORBENS", "compost"])

    # Operating conditions as a one-row metadata file.
    write_csv(
        OUT / "operating_conditions.csv",
        [
            {
                "Q_L_min": Q_LMIN,
                "Q_m3_s": Q_M3S,
                "diameter_m": DIAM_M,
                "area_m2": AREA_M2,
                "bed_length_m": BED_LEN_M,
                "ebct_s": AREA_M2 * BED_LEN_M / Q_M3S,
            }
        ],
        ["Q_L_min", "Q_m3_s", "diameter_m", "area_m2", "bed_length_m", "ebct_s"],
    )

    # Quick integrity summary.
    runs = sorted({row["run_date"] for row in profiles})
    print(f"\n{len(profiles)} measurements across {len(runs)} runs: {runs}")


if __name__ == "__main__":
    main()
